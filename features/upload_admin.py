"""Data admin and upload feature module.

V6.5 moves the data-upload/admin implementation out of core.legacy.
The code is intentionally behavior-preserving: shared UI/data helpers still come
from core.legacy while upload-specific helpers live here.
"""

from __future__ import annotations

import os
import re
import sys
import tempfile
import zipfile
import gc
import time
from typing import Optional

import pandas as pd
import streamlit as st
from google.cloud import bigquery

from config import BQ_PREFIX
from services.data_cleaner import clean_upload_dataframe
from services.bigquery_client import get_bq_client
from services.etl_refresh import refresh_core_marts, format_refresh_results
from core.legacy import (
    hero,
    section_header,
    to_datetime_safe,
    load_daily_ops,
    _recent_month_labels,
    parse_winback_file,
    _winback_ym_from_name,
    _bq_winback_month_exists,
    _write_winback,
)

def _import_tool():
    """惰性载入同目录的 import_tool（部署时与 dashboard.py 一起进 repo）。"""
    import os
    import sys
    here = os.path.dirname(os.path.abspath(__file__))
    if here not in sys.path:
        sys.path.insert(0, here)
    import import_tool
    return import_tool


def _bq_count_source_file(client, table_name: str, source_file: str) -> int:
    """既有表里这个 _source_file 已写过几行（防同档重复）。表不存在/无此栏 → 0。"""
    sql = (
        f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table_name}` "
        f"WHERE _source_file = @sf"
    )
    cfg = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter('sf', 'STRING', source_file)
    ])
    try:
        return list(client.query(sql, job_config=cfg).result())[0].n
    except Exception:
        return 0


def _identify_by_content(it, tmp_path):
    """档名认不出时，靠栏位内容认报表类型（裸档名 top.xlsx、或重命名过的档都能认）。
    只用各报表「最有辨识度」的栏位组合，避免误判。"""
    try:
        cols = set(str(c).strip() for c in it.read_file(tmp_path).columns)
    except Exception:
        return None

    def has(*xs):
        return all(x in cols for x in xs)
    if has('用户名', '个人输赢') or has('用户名', '杀数'):
        return ('raw_top_report', 'Top报表')
    if has('会员账号', '注册时间', '是否为代理'):
        return ('raw_member_report', '会员报表')
    if has('注册数', '公司输赢', '首存人数'):
        return ('raw_platform_report', '平台报表')
    if has('红利', '返水', '代理佣金'):
        return ('raw_finance_report', '财务报表')
    return None


def _parse_standard_report(it, tmp_path: str, source_file: str, table_override=None):
    """复用 import_tool 解析单个标准月报。返回 (table, display, df_data)；无法识别→(None,None,None)。
    table_override=(table, display)：档名认不出时由内容识别传入。"""
    table, display = it.identify_report_type(source_file)
    if table is None and table_override:
        table, display = table_override
    if table is None:
        return None, None, None
    df = it.read_file(tmp_path)
    summary_mask = df.apply(lambda r: it.is_summary_row(r.values), axis=1)
    df_clean = df[~summary_mask].copy()
    zero_mask = df_clean.apply(
        lambda r: it.is_all_zero_data_row(r.values, df_clean.columns), axis=1)
    df_data = df_clean[~zero_mask].copy()
    # 快照月份（TOP / 会员）
    if table == 'raw_top_report':
        month = it.extract_top_month(source_file)
        if month:
            df_data['_snapshot_month'] = month
            snap = it.month_to_snapshot_date(month)
            if snap:
                df_data['_snapshot_date'] = snap
    elif table == 'raw_member_report':
        month = it.infer_member_snapshot_month(df_data, tmp_path)
        if month:
            df_data['_snapshot_month'] = month
            snap = it.month_to_snapshot_date(month)
            if snap:
                df_data['_snapshot_date'] = snap
    return table, display, df_data


def _append_standard_report(client, df_data, table_name: str, source_file: str) -> int:
    """append 写入（绝不覆盖）。沿用 import_tool 的 schema 容错。"""
    payload = clean_upload_dataframe(df_data.copy())
    payload['_imported_at'] = pd.Timestamp.now()
    payload['_source_file'] = source_file
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION],
    )
    job = client.load_table_from_dataframe(
        payload, f"{BQ_PREFIX}.{table_name}", job_config=cfg)
    job.result()
    return job.output_rows


# ---- 红利 / 客服对话（第二批，全部 append-only，绝不 TRUNCATE，不依赖 DML）----

def _parse_bonus_df(it, csv_path):
    """单档红利清洗（复用 import_bonus_records 逻辑）。返回 df（含 订单号/红利金额/_snapshot_month/活动名称）。"""
    import pandas as pd
    df = None
    for enc in ('utf-8', 'gb18030', 'gbk', 'big5'):
        try:
            df = pd.read_csv(csv_path, encoding=enc)
            break
        except Exception:
            continue
    if df is None:
        raise ValueError('红利 CSV 编码识别失败')
    if '流水倍数(倍)' in df.columns:
        df = df.rename(columns={'流水倍数(倍)': '流水倍数'})
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = (df[col].astype(str)
                       .str.replace(r'^="(.*)"$', r'\1', regex=True).str.strip())
            df[col] = df[col].replace({'nan': None, 'None': None, '': None})
    if '红利标题' in df.columns:
        df['活动名称'] = df['红利标题'].fillna('').replace('', None)
        mask = df['活动名称'].isna()
        if '申请备注' in df.columns:
            df.loc[mask, '活动名称'] = df.loc[mask, '申请备注']
        df['活动名称'] = df['活动名称'].fillna('未知')
    if '申请时间' in df.columns:
        dt = pd.to_datetime(df['申请时间'], errors='coerce')
        df['_snapshot_month'] = dt.dt.strftime('%Y-%m')
    if '订单号' in df.columns:
        df = df.drop_duplicates(subset=['订单号'], keep='first')
    for col in ('红利金额', '流水倍数'):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def _parse_cs_df(it, xlsx_path, source_file):
    """单档客服对话清洗（复用 import_cs_conversations 逻辑，多 sheet）。月份从 开始时间 推断。"""
    import re as _re
    import pandas as pd
    xl = pd.ExcelFile(xlsx_path)
    frames = []
    for sh in [s for s in xl.sheet_names if s != '总表']:
        try:
            d = pd.read_excel(xlsx_path, sheet_name=sh)
        except Exception:
            continue
        if d.empty:
            continue
        d['_sheet'] = sh
        frames.append(d)
    if not frames:
        raise ValueError('客服对话 xlsx 没读到数据 sheet')
    merged = pd.concat(frames, ignore_index=True)
    for c in ('开始时间', '结束时间'):
        if c in merged.columns:
            merged[c] = pd.to_datetime(merged[c], errors='coerce')
    month = ''
    if '开始时间' in merged.columns and merged['开始时间'].notna().any():
        mode = merged['开始时间'].dt.strftime('%Y%m').mode()
        if len(mode):
            month = mode.iloc[0]
    if not month:
        m = _re.search(r'(20\d{2})\D?([01]\d)', source_file)
        month = f'{m.group(1)}{m.group(2)}' if m else ''
    merged['_snapshot_month'] = month
    for c in ('首次响应', '平均响应', '总时长', '访客消息数', '客服消息数', '撤回消息数', '对话回合数'):
        if c in merged.columns:
            merged[c] = pd.to_numeric(merged[c], errors='coerce')
    str_cols = ['终端', '访客ID', '对话ID', '新对话ID', '会员账号', '地区', '接待客服',
                '访客IP', '网站名称', '是否邀请评价', '满意度评价', '评价内容',
                '服务主题', '备注', '机器人标识', '对话内容', '_sheet']
    for c in str_cols:
        if c in merged.columns:
            merged[c] = merged[c].astype(str).replace({'nan': None, 'None': None, '': None})
    UNHAPPY = {'非常不满意', '不满意'}

    def _ext(row):
        if str(row.get('满意度评价') or '') not in UNHAPPY:
            return ''
        return it.extract_unhappy_reason(row)
    merged['_extracted_issue'] = [_ext(merged.iloc[i]) for i in range(len(merged))]
    return merged


def _existing_bonus_orders(client):
    try:
        rows = client.query(
            f"SELECT `订单号` FROM `{BQ_PREFIX}.raw_bonus_report`").result()
        return set(str(r['订单号']) for r in rows)
    except Exception:
        return set()


def _append_bonus(client, df, source_file, existing_orders=None):
    """只追加「新订单号」的红利行，既有订单一律不动。返回真正写入行数。"""
    import pandas as pd
    if existing_orders is None:
        existing_orders = _existing_bonus_orders(client)
    if '订单号' in df.columns and existing_orders:
        new = df[~df['订单号'].astype(str).isin(existing_orders)].copy()
    else:
        new = df.copy()
    if len(new) == 0:
        return 0
    new['_source_file'] = source_file
    new['_imported_at'] = pd.Timestamp.now()
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION])
    client.load_table_from_dataframe(
        new, f"{BQ_PREFIX}.raw_bonus_report", job_config=cfg).result()
    return len(new)


def _cs_basename_loaded(client, basename: str) -> int:
    sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.raw_cs_conversations` "
           f"WHERE _source_file = @b OR ENDS_WITH(_source_file, @sb)")
    cfg = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter('b', 'STRING', basename),
        bigquery.ScalarQueryParameter('sb', 'STRING', '/' + basename)])
    try:
        return list(client.query(sql, job_config=cfg).result())[0].n
    except Exception:
        return 0


def _append_cs(client, df, source_file):
    import pandas as pd
    payload = df.copy()
    payload['_source_file'] = source_file
    payload['_imported_at'] = pd.Timestamp.now()
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION])
    client.load_table_from_dataframe(
        payload, f"{BQ_PREFIX}.raw_cs_conversations", job_config=cfg).result()
    return len(payload)


# ---- 代理佣金 单线/团队版（第三批，append-safe by 佣金月份；绝不依赖 DML）----
_COMM_NUMERIC = {'存款金额', '提款金额', '总输赢', '场馆费', '红利', '代理冲销', '返水', '账户调整',
    '存款手续费基数', '提款手续费基数', '存款手续费', '提款手续费', '补单输赢', '净输赢', '上月结余',
    '冲账调整', '冲正后净输赢', '佣金调整', '佣金', '已发放佣金', '剩余佣金', '申请发放佣金', 'VIP专享'}
_COMM_PERCENT = {'佣金比例', '二次佣金比例'}
_COMM_INT = {'团队人数', '下级人数', '注册人数', '首存人数', '活跃人数', '新增活跃人数', '有效活跃人数'}
_COMM_DATE = {'成为代理时间', '加入团队时间'}
_COMM_BOOL = {'是否在团队', '是否取消代理资格', '是否为主线'}


def _comm_num(v):
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().replace(',', '')
    if s in ('', '-', 'None', 'nan'):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _comm_pct(v):
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().rstrip('%')
    if s in ('', '-', 'None', 'nan'):
        return None
    try:
        return float(s) / 100.0
    except Exception:
        return None


def _comm_int(v):
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def _comm_bool(v):
    s = str(v).strip() if v is not None else ''
    return True if s == '是' else (False if s == '否' else None)


def _commission_table(cols):
    cset = set(str(c).strip() for c in cols)
    if ('线别' in cset or '主线/副线' in cset) and '团队名称' in cset:
        return 'raw_agent_commission_team'
    if '是否在团队' in cset:
        return 'raw_agent_commission_single'
    return None


def _parse_commission(it, path):
    import pandas as pd
    df = it.read_file(path)
    if '主线/副线' in df.columns:
        df = df.rename(columns={'主线/副线': '线别'})
    for c in df.columns:
        cs = str(c).strip()
        if cs in _COMM_NUMERIC:
            df[c] = df[c].map(_comm_num)
        elif cs in _COMM_PERCENT:
            df[c] = df[c].map(_comm_pct)
        elif cs in _COMM_INT:
            df[c] = df[c].map(_comm_int)
        elif cs in _COMM_DATE:
            df[c] = pd.to_datetime(df[c], errors='coerce')
        elif cs in _COMM_BOOL:
            df[c] = df[c].map(_comm_bool)
        else:
            df[c] = df[c].map(lambda v: None if v is None or str(v).strip() in ('', 'nan', 'None') else str(v).strip())
    keep = [c for c in df.columns if c not in ('_source_file', '_imported_at')]
    return df.dropna(subset=keep, how='all')


def _commission_months(df):
    if '佣金月份' in df.columns:
        return sorted(set(df['佣金月份'].dropna().astype(str)))
    return []


def _commission_month_exists(client, table, months):
    if not months:
        return 0
    sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table}` "
           f"WHERE CAST(`佣金月份` AS STRING) IN UNNEST(@ms)")
    cfg = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ArrayQueryParameter('ms', 'STRING', months)])
    try:
        return list(client.query(sql, job_config=cfg).result())[0].n
    except Exception:
        return 0


def _write_commission_safe(client, new_df, table, source_files):
    """读现有→去掉 new 的月份→concat→统一类型→TRUNCATE 写回。
    保留其他月份不动、刷新上传的月份；写回前校验「其他月份行数不变」防掉数据。返回写入总行数。"""
    import pandas as pd
    new_months = set(new_df['佣金月份'].dropna().astype(str)) if '佣金月份' in new_df.columns else set()
    try:
        existing = client.query(f"SELECT * FROM `{BQ_PREFIX}.{table}`").result().to_dataframe()
    except Exception:
        # 第一次导入时表不存在，先当作空表，再由 load_table_from_dataframe 自动建表。
        existing = pd.DataFrame()
    if '佣金月份' in existing.columns and new_months:
        keep = existing[~existing['佣金月份'].astype(str).isin(new_months)]
    else:
        keep = existing
    nd = new_df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_files
    combined = pd.concat([keep, nd], ignore_index=True)
    for c in list(combined.columns):
        cs = str(c).strip()
        if cs in (_COMM_NUMERIC | _COMM_PERCENT | _COMM_INT):
            combined[c] = pd.to_numeric(combined[c], errors='coerce')
        elif cs in _COMM_DATE:
            combined[c] = pd.to_datetime(combined[c].astype(str), errors='coerce')
        elif cs in _COMM_BOOL:
            combined[c] = combined[c].map(
                lambda v: True if v in (True, 'True', '是') else (False if v in (False, 'False', '否') else None)
            ).astype('boolean')
    # 防掉数据：每个「没在上传的旧月份」行数必须原样保留
    if '佣金月份' in existing.columns:
        ex_counts = existing['佣金月份'].astype(str).value_counts().to_dict()
        cb_counts = combined['佣金月份'].astype(str).value_counts().to_dict()
        for m, cnt in ex_counts.items():
            if m not in new_months and cb_counts.get(m, 0) != cnt:
                raise RuntimeError(f'安全中止：旧月份 {m} 行数会变（{cnt}→{cb_counts.get(m, 0)}），拒绝写入')
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(combined, f"{BQ_PREFIX}.{table}", job_config=cfg).result()
    return len(combined)


# ── 按月快照报表（会员 / TOP）同月替换 + 删除月份（read-modify-truncate，免费版真删）──

SNAPSHOT_TABLES = {'raw_member_report', 'raw_top_report'}  # 有 _snapshot_month、同月应替换


def _replace_by_snapshot_month(client, new_df, table, source_file):
    """会员/TOP 等按月快照表：同月「按会员账号+代理合并」（不是整月覆盖）。
    同一个月里：上传档有的(账号+代理)→用新的更新；只在旧档有的→保留；新的→加进来；其他月份完全不动。
    去重必须含代理：同名挂不同代理=不同人，只按账号会把另一个代理下的同名会员误删（跟 member_count 一致）。
    这样分批传同一个月（如 6/1-12 再传 6/13-30）会累加合并、不丢数据。
    返回 (months_str, updated, added, total)。"""
    months = (set(new_df['_snapshot_month'].dropna().astype(str))
              if '_snapshot_month' in new_df.columns else set())
    try:
        existing = client.query(f"SELECT * FROM `{BQ_PREFIX}.{table}`").result().to_dataframe()
    except Exception:
        # 第一次导入时表不存在，先当作空表，再由 load_table_from_dataframe 自动建表。
        existing = pd.DataFrame()
    has_acct = '会员账号' in existing.columns and '会员账号' in new_df.columns
    has_agent = has_acct and '代理' in existing.columns and '代理' in new_df.columns

    def _merge_key(d):  # 账号+代理；没代理列才退回纯账号
        if has_agent:
            return d['会员账号'].astype(str) + '\x01' + d['代理'].astype(str)
        return d['会员账号'].astype(str)

    new_keys = set(_merge_key(new_df)) if has_acct else set()
    if '_snapshot_month' in existing.columns and months and has_acct:
        # 去掉「同月 且 (账号+代理)在新档里」的旧行（被新行取代）；同月没在新档的 + 其他月，全保留
        same_month = existing['_snapshot_month'].astype(str).isin(months)
        old_mask = same_month & _merge_key(existing).isin(new_keys)
        updated = int(old_mask.sum())
        kept_same_month = int((same_month & ~old_mask).sum())  # 同月、新档没有 → 保留的旧会员
        added = len(new_df) - updated
    elif '_snapshot_month' in existing.columns and months:
        # 没有会员账号栏（如某些 TOP）→ 退回整月替换
        old_mask = existing['_snapshot_month'].astype(str).isin(months)
        updated = int(old_mask.sum())
        added = len(new_df)
    else:
        old_mask = pd.Series(False, index=existing.index)
        updated = 0
        added = len(new_df)
    keep = existing[~old_mask].copy()
    nd = new_df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_file
    combined = pd.concat([keep, nd], ignore_index=True)
    # 防掉数据：没在上传的「其他月份」行数必须原样保留
    if '_snapshot_month' in existing.columns:
        ex_counts = existing['_snapshot_month'].astype(str).value_counts().to_dict()
        cb_counts = combined['_snapshot_month'].astype(str).value_counts().to_dict()
        for m, cnt in ex_counts.items():
            if m not in months and cb_counts.get(m, 0) != cnt:
                raise RuntimeError(f'安全中止：其他月份 {m} 行数会变（{cnt}→{cb_counts.get(m, 0)}），拒绝写入')
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(combined, f"{BQ_PREFIX}.{table}", job_config=cfg).result()
    return ','.join(sorted(months)) if months else '?', updated, added, len(combined)


def _date_col_of(df) -> Optional[str]:
    """日报类报表的日期栏：日期 或 时间（取前 10 字 = YYYY-MM-DD，realtime「..23~24」也 OK）。"""
    if '日期' in df.columns:
        return '日期'
    if '时间' in df.columns:
        return '时间'
    if '订单时间' in df.columns:   # 存款订单
        return '订单时间'
    if '申请时间' in df.columns:   # 提款订单
        return '申请时间'
    return None


def _date_keys(series) -> set:
    ks = set(series.astype(str).str[:10])
    ks.discard('')
    ks.discard('nan')
    ks.discard('None')
    return ks


def _looks_month_aggregated(df) -> bool:
    """日报类报表若日期栏大多只有 YYYY-MM（没有「日」），多半是后台误用「按月」颗粒度导出的
    「整月汇总」而非「日明细」。导进去后按日期范围筛会整月对不上（跟其他月日明细不一致）。"""
    dc = _date_col_of(df)
    if dc is None:
        return False
    s = df[dc].dropna().astype(str).str.strip()
    s = s[(s != '') & (s.str.lower() != 'nan') & (s != 'none')]
    if s.empty:
        return False
    month_only = s.str.match(r'^\d{4}-\d{2}$')
    return float(month_only.mean()) > 0.8


def _replace_by_date_range(client, new_df, table, source_file):
    """日报类（平台/财务/游戏/场馆/分析/推广/代理/实时注单）：按日期替换。
    去掉「现有表里日期在新档日期集合内」的旧行→写新行，其它日期一行不动。
    重传同一天/同月 = 覆盖那些日期，永不重复。

    第一次导入时 BigQuery 表可能还不存在。
    原本这里会先 SELECT 现有表，导致 404 Not found。
    现在改成：表不存在时 existing = 空 DataFrame，再由 load_table_from_dataframe 自动建表。
    返回 (date_range_str, removed, written, total)。
    """
    dc = _date_col_of(new_df)
    if dc is None:
        # 没日期栏 → 退回按档名追加（理论上日报类都有日期，不该走到这）
        n = _append_standard_report(client, new_df, table, source_file)
        return '(无日期栏)', 0, n, None

    new_df = clean_upload_dataframe(new_df.copy())
    new_dates = _date_keys(new_df[dc])
    new_months = {d[:7] for d in new_dates}

    try:
        existing = client.query(f"SELECT * FROM `{BQ_PREFIX}.{table}`").result().to_dataframe()
    except Exception:
        # 第一次导入时表不存在，直接从空表开始。
        existing = pd.DataFrame()

    if dc in existing.columns and new_dates:
        ex_key = existing[dc].astype(str).str[:10]
        # 覆盖同一天的旧行；并顺手清掉「同月的整月汇总行」(如旧的 2026-05)，
        # 否则它的 key 不等于任何 2026-05-DD、会跟新的日明细并存造成重复/口径混乱
        stale_month_agg = ex_key.str.match(r'^\d{4}-\d{2}$') & ex_key.str[:7].isin(new_months)
        old_mask = ex_key.isin(new_dates) | stale_month_agg
    else:
        old_mask = pd.Series(False, index=existing.index)

    removed = int(old_mask.sum())
    keep = existing[~old_mask].copy()

    # 防掉数据：保留的行数必须 = 现有 - 删除
    if len(keep) != len(existing) - removed:
        raise RuntimeError('安全中止：行数对不上，拒绝写入')

    nd = new_df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_file

    combined = pd.concat([keep, nd], ignore_index=True)

    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
        autodetect=True,
    )
    client.load_table_from_dataframe(combined, f"{BQ_PREFIX}.{table}", job_config=cfg).result()

    rng = (f"{min(new_dates)} ~ {max(new_dates)}" if len(new_dates) > 1
           else (next(iter(new_dates)) if new_dates else '?'))
    return rng, removed, len(nd), len(combined)


def _table_period_col(table: str):
    """这张表用哪个栏当「月份/期间」键。没有→None（只能按上传档名删）。"""
    if table in SNAPSHOT_TABLES:
        return '_snapshot_month'
    if table in ('raw_agent_commission_single', 'raw_agent_commission_team'):
        return '佣金月份'
    if table in ('raw_winback', 'raw_agent_settlement_monthly'):
        return '月份'
    return None


def _bq_periods(client, table: str):
    """回传 (period_col, {期间: 行数})。有月份栏就按月份；否则按 _source_file（上传批次）。"""
    col = _table_period_col(table)
    key = col if col else '_source_file'
    try:
        sql = (f"SELECT CAST(`{key}` AS STRING) AS p, COUNT(*) AS n "
               f"FROM `{BQ_PREFIX}.{table}` GROUP BY p ORDER BY p")
        rows = list(client.query(sql).result())
        return key, {(r.p if r.p is not None else '(空)'): r.n for r in rows}
    except Exception:
        return key, {}


def _bq_delete_periods(client, table: str, key: str, periods):
    """真删：读现有→去掉指定期间→TRUNCATE 写回。返回 (removed, remaining)。"""
    periods = set(str(p) for p in periods)
    try:
        existing = client.query(f"SELECT * FROM `{BQ_PREFIX}.{table}`").result().to_dataframe()
    except Exception:
        return 0, 0
    if key not in existing.columns:
        return 0, len(existing)
    mask = existing[key].astype(str).isin(periods)
    removed = int(mask.sum())
    keep = existing[~mask].copy()
    if len(keep) != len(existing) - removed:
        raise RuntimeError('安全中止：行数对不上，拒绝写入')
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(keep, f"{BQ_PREFIX}.{table}", job_config=cfg).result()
    return removed, len(keep)


def _try_parse_agent_monthly(tmp_path):
    """市代「整理资料」分页：每代理每月 充值/投注/输赢/挂账。返回 (df, months) 或 None。
    认法：xlsx 里有一张分页含「代理帐号 + 月份 + (累计挂账额度 或 发放佣金)」栏。"""
    import os
    if os.path.splitext(tmp_path)[1].lower() not in ('.xlsx', '.xls'):
        return None
    try:
        xl = pd.ExcelFile(tmp_path)
    except Exception:
        return None
    target = None
    for sh in xl.sheet_names:
        try:
            cols = set(str(c).strip() for c in pd.read_excel(tmp_path, sheet_name=sh, nrows=0).columns)
        except Exception:
            continue
        if {'代理帐号', '月份'}.issubset(cols) and ('累计挂账额度' in cols or '发放佣金' in cols):
            target = sh
            break
    if target is None:
        return None
    df = pd.read_excel(tmp_path, sheet_name=target, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['代理帐号'].notna() & (df['代理帐号'].astype(str).str.strip() != '')].copy()
    mm = pd.to_datetime(df['月份'], errors='coerce')
    df = df[mm.notna()].copy()
    if df.empty:
        return None
    df['月份'] = pd.to_datetime(df['月份'], errors='coerce').dt.strftime('%Y-%m')
    text_cols = {'代理名称', '代理帐号', '发展情况', '开户日期', '月份', '来源栏位'}
    for c in df.columns:
        if c in text_cols:
            df[c] = df[c].astype(str).str.strip()
        else:
            df[c] = pd.to_numeric(df[c], errors='coerce')
    months = sorted(df['月份'].dropna().unique().tolist())
    return df, months


def _write_agent_monthly(client, df, months, source_file):
    """写 raw_agent_settlement_monthly：按月份替换（读改 WRITE_TRUNCATE，沙盒禁 DELETE）。返回写入行数。"""
    table = 'raw_agent_settlement_monthly'
    nd = df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_file
    full = f'{BQ_PREFIX}.{table}'
    try:
        existing = client.query(f'SELECT * FROM `{full}`').result().to_dataframe()
    except Exception:
        existing = pd.DataFrame()
    if not existing.empty and '月份' in existing.columns:
        keep = existing[~existing['月份'].astype(str).isin([str(m) for m in months])].copy()
        combined = pd.concat([keep, nd], ignore_index=True)
    else:
        combined = nd
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(combined, full, job_config=cfg).result()
    return len(nd)


def _try_parse_agent_settlement(tmp_path):
    """平哥「X月代理帐.xlsx」（位置式：R1/R2 累计挂帐摘要 + R4/R5 发放摘要 + R9+ 明细 9 栏）。
    返回 (summary_records, detail_records) 或 None。月份在档外，写入时再补。"""
    import os
    if os.path.splitext(tmp_path)[1].lower() not in ('.xlsx', '.xls'):
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path, data_only=True)
    except Exception:
        return None
    sheet = next((s for s in wb.sheetnames if ('代理帐' in s or '代理账' in s)), wb.sheetnames[0])
    rows = list(wb[sheet].iter_rows(values_only=True))
    if len(rows) < 10:
        return None
    r7 = [str(c).strip() if c is not None else '' for c in (rows[7] or [])]
    if '名称' not in r7 or '总代' not in r7:   # 不是这种位置式格式
        return None

    def _num(v):
        return float(v) if isinstance(v, (int, float)) else None
    summary = []
    if len(rows) > 2:
        l1, v1 = rows[1] or [], rows[2] or []
        for i in range(1, 7):
            if i < len(l1) and l1[i] and i < len(v1) and isinstance(v1[i], (int, float)):
                summary.append({'项目': str(l1[i]).strip(), '金额': float(v1[i])})
    if len(rows) > 5:
        l2, v2 = rows[4] or [], rows[5] or []
        for i in range(3, 7):
            if i < len(l2) and l2[i] and i < len(v2) and isinstance(v2[i], (int, float)):
                summary.append({'项目': str(l2[i]).strip(), '金额': float(v2[i])})
    detail = []
    for r in rows[8:]:
        if not any(v not in (None, '') for v in r):
            continue
        treatment = r[6] if len(r) > 6 else None
        name = r[1] if len(r) > 1 else None
        agent = r[2] if len(r) > 2 else None
        if not treatment or not name or not agent:
            continue
        detail.append({
            '回款状态': str(r[0]).strip() if r[0] else None,
            '名称': str(name).strip(), '总代账号': str(agent).strip(),
            '先前挂账业绩': _num(r[3]) if len(r) > 3 else None,
            '本月业绩': _num(r[4]) if len(r) > 4 else None,
            '业绩总计': _num(r[5]) if len(r) > 5 else None,
            '适用待遇': str(treatment).strip(),
            '比例': _num(r[7]) if len(r) > 7 else None,
            '实际佣金': _num(r[8]) if len(r) > 8 else None,
        })
    if not summary and not detail:
        return None
    return summary, detail


def _write_agent_settlement(client, summary_records, detail_records, month, source_file=''):
    """写 raw_agent_settlement_summary + _detail：按月份替换（读改 WRITE_TRUNCATE）。返回 (摘要行, 明细行)。"""
    def _write_one(table, records):
        if not records:
            return 0
        df = pd.DataFrame(records)
        df['月份'] = month
        df['_imported_at'] = pd.Timestamp.now().isoformat()   # 跟原表 _imported_at 同为字符串，避免 Timestamp 类型冲突
        df['_source_file'] = source_file or f'代理帐_{month}.xlsx'
        full = f'{BQ_PREFIX}.{table}'
        try:
            existing = client.query(f'SELECT * FROM `{full}`').result().to_dataframe()
        except Exception:
            existing = pd.DataFrame()
        if not existing.empty and '月份' in existing.columns:
            keep = existing[existing['月份'].astype(str) != str(month)].copy()
            combined = pd.concat([keep, df], ignore_index=True)
        else:
            combined = df
        cfg = bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
        client.load_table_from_dataframe(combined, full, job_config=cfg).result()
        return len(df)
    ns = _write_one('raw_agent_settlement_summary', summary_records)
    nd = _write_one('raw_agent_settlement_detail', detail_records)
    return ns, nd


def _classify_and_parse(it, client, src, tmp_path):
    """识别+解析一个上传单元。返回 entry dict（kind/display/df/rows/is_new/status）。"""
    import os
    # 1) 标准月报（先文件名匹配 FILE_MAP；认不出再靠栏位内容认）
    table, display = it.identify_report_type(src)
    content_override = None
    if table is None:
        content_override = _identify_by_content(it, tmp_path)
        if content_override:
            table, display = content_override
    if table is not None:
        _, disp, df = _parse_standard_report(it, tmp_path, src, table_override=content_override)
        # 会员 / TOP 是按月快照：同月「按会员账号合并」（分批传同月会累加、不丢；换档名也不会变两份）
        if table in SNAPSHOT_TABLES:
            mth = (str(df['_snapshot_month'].iloc[0])
                   if '_snapshot_month' in df.columns and len(df) else '')
            if mth in ('', 'None', 'nan', '?'):
                # 抓不到月份（如 TOP 报表 top.xlsx 裸档名、内容无日期栏）→ 让上传页选月份再写
                return {'src': src, 'kind': 'snapshot', 'table': table, 'display': disp,
                        'df': df, 'rows': len(df), 'is_new': len(df) > 0,
                        'snapshot_month': None, 'need_month': True, 'exists': 0,
                        'status': '🗓 需选月份（档名/内容判不出是哪个月，请在下方选）'}
            try:
                sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table}` "
                       f"WHERE CAST(`_snapshot_month` AS STRING)=@m")
                cfg = bigquery.QueryJobConfig(query_parameters=[
                    bigquery.ScalarQueryParameter('m', 'STRING', mth)])
                exists = list(client.query(sql, job_config=cfg).result())[0].n
            except Exception:
                exists = 0
            return {'src': src, 'kind': 'snapshot', 'table': table, 'display': disp,
                    'df': df, 'rows': len(df), 'is_new': len(df) > 0, 'snapshot_month': mth,
                    'exists': exists,
                    'status': (f'🔄 合并到 {mth}（同月已有 {exists} 行，按会员账号合并：重复的更新、新会员追加、其他月不动）'
                               if exists else f'🆕 新增 {mth}')}
        # 日报类（有日期/时间栏）：按日期替换——重传同一天/同月就覆盖那些日期，永不重复
        dc = _date_col_of(df)
        if dc:
            nd_dates = _date_keys(df[dc])
            rng = (f"{min(nd_dates)} ~ {max(nd_dates)}" if len(nd_dates) > 1
                   else (next(iter(nd_dates)) if nd_dates else '?'))
            try:
                sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table}` "
                       f"WHERE SUBSTR(CAST(`{dc}` AS STRING),1,10) IN UNNEST(@d)")
                cfg = bigquery.QueryJobConfig(query_parameters=[
                    bigquery.ArrayQueryParameter('d', 'STRING', list(nd_dates))])
                exists = list(client.query(sql, job_config=cfg).result())[0].n
            except Exception:
                exists = 0
            month_agg = _looks_month_aggregated(df)
            base_status = (f'🔄 替换 {rng}（覆盖这段日期已有的 {exists} 行，其他日期不动）'
                           if exists else f'🆕 新增 {rng}')
            if month_agg:
                base_status = '⚠️ 疑似「整月汇总」非日明细（日期只有年月、没到日）— ' + base_status
            return {'src': src, 'kind': 'standard', 'table': table, 'display': disp,
                    'df': df, 'rows': len(df), 'is_new': len(df) > 0, 'date_range': rng,
                    'date_col': dc, 'date_keys': sorted(nd_dates), 'exists': exists,
                    'warn_month_agg': month_agg, 'status': base_status}
        # 没日期/时间栏（理论上日报类都有）→ 退回按档名去重
        dup = _bq_count_source_file(client, table, src)
        return {'src': src, 'kind': 'standard', 'table': table, 'display': disp,
                'df': df, 'rows': len(df), 'is_new': dup == 0,
                'status': (f'⏭ 已上传过（{dup} 行），会跳过' if dup else '🆕 待写入')}
    # 1.34) 代理结算月报（平哥「X月代理帐.xlsx」位置式：累计挂帐摘要 + 适用待遇/退成明细）→ 需选月份、按月替换
    sett = _try_parse_agent_settlement(tmp_path)
    if sett is not None:
        summ, det = sett
        return {'src': src, 'kind': 'settlement', 'display': '代理结算月报(平哥)',
                'summary_records': summ, 'detail_records': det, 'df': None,
                'rows': len(det), 'is_new': (len(summ) + len(det)) > 0,
                'need_month': True, 'snapshot_month': None,
                'status': f'🗓 需选月份（代理结算月报：{len(summ)} 摘要 + {len(det)} 明细）'}
    # 1.35) 代理结算月度（市代「整理资料」分页：每代理每月 充值/投注/输赢/挂账）→ 按月份替换
    am = _try_parse_agent_monthly(tmp_path)
    if am is not None:
        am_df, am_months = am
        try:
            sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.raw_agent_settlement_monthly` "
                   f"WHERE CAST(`月份` AS STRING) IN UNNEST(@m)")
            cfg = bigquery.QueryJobConfig(query_parameters=[
                bigquery.ArrayQueryParameter('m', 'STRING', am_months)])
            exists = list(client.query(sql, job_config=cfg).result())[0].n
        except Exception:
            exists = 0
        rng = f'{am_months[0]} ~ {am_months[-1]}' if len(am_months) > 1 else (am_months[0] if am_months else '?')
        return {'src': src, 'kind': 'agent_monthly', 'table': 'raw_agent_settlement_monthly',
                'display': '代理结算月度(市代)', 'df': am_df, 'rows': len(am_df),
                'is_new': len(am_df) > 0, 'months': am_months,
                'status': (f'🔄 按月替换 {rng}（{len(am_months)} 个月，覆盖同月已有 {exists} 行，其他月不动）'
                           if exists else f'🆕 新增 {rng}（{len(am_months)} 个月）')}
    # 1.4) 存取款订单（存款管理/提款管理 历史记录 Csv）→ 按日期替换
    if ('存款管理' in src) or ('提款管理' in src):
        is_dep = '存款管理' in src
        table = 'raw_finance_deposit' if is_dep else 'raw_finance_withdraw'
        disp = '存款订单' if is_dep else '提款订单'
        fdf = None
        for enc in ('utf-8-sig', 'gbk', 'utf-8', 'big5'):
            try:
                fdf = pd.read_csv(tmp_path, dtype=str, encoding=enc)
                break
            except Exception:
                continue
        if fdf is None or fdf.empty:
            return {'src': src, 'kind': 'none', 'display': disp, 'df': None, 'rows': 0,
                    'is_new': False, 'status': '⚠️ 读不出该 Csv'}
        for c in fdf.columns:
            fdf[c] = fdf[c].astype(str).str.replace('\t', '', regex=False).str.strip()
        # 去掉栏名里的括号单位（如 订单金额(元)→订单金额）；BigQuery 栏名不允许括号
        fdf.columns = [re.sub(r'[（(].*?[）)]', '', str(c)).strip() for c in fdf.columns]
        dc = '订单时间' if is_dep else '申请时间'
        if dc not in fdf.columns:
            return {'src': src, 'kind': 'none', 'display': disp, 'df': None, 'rows': 0,
                    'is_new': False, 'status': f'⚠️ 缺日期栏「{dc}」，认不出'}
        nd_dates = _date_keys(fdf[dc])
        rng = (f"{min(nd_dates)} ~ {max(nd_dates)}" if len(nd_dates) > 1
               else (next(iter(nd_dates)) if nd_dates else '?'))
        try:
            sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table}` "
                   f"WHERE SUBSTR(CAST(`{dc}` AS STRING),1,10) IN UNNEST(@d)")
            cfg = bigquery.QueryJobConfig(query_parameters=[
                bigquery.ArrayQueryParameter('d', 'STRING', list(nd_dates))])
            exists = list(client.query(sql, job_config=cfg).result())[0].n
        except Exception:
            exists = 0
        return {'src': src, 'kind': 'standard', 'table': table, 'display': disp, 'df': fdf,
                'rows': len(fdf), 'is_new': len(fdf) > 0, 'date_range': rng,
                'date_col': dc, 'date_keys': sorted(nd_dates), 'exists': exists,
                'status': (f'🔄 替换 {rng}（覆盖这段日期已有的 {exists} 行）' if exists else f'🆕 新增 {rng}')}

    # 1.5) 电访召回「撥打紀錄總表」→ 解析成各专员月度汇总，存 raw_winback（按月份刷新）
    if ('撥打' in src) or ('拨打' in src) or ('電訪' in src) or ('电访' in src):
        try:
            wdf, wmeta = parse_winback_file(tmp_path)
        except Exception as e:
            return {'src': src, 'kind': 'none', 'display': '电访召回(解析失败)', 'df': None,
                    'rows': 0, 'is_new': False, 'status': f'⚠️ 撥打紀錄總表解析失败：{str(e)[:50]}'}
        ym = wmeta.get('month') or _winback_ym_from_name(src)
        if wdf is None or wdf.empty or not ym:
            return {'src': src, 'kind': 'none', 'display': '电访召回', 'df': None, 'rows': 0,
                    'is_new': False, 'status': '⚠️ 没读到专员数据或判不出月份，跳过'}
        wdf = wdf.copy()
        wdf['月份'] = ym
        exists = _bq_winback_month_exists(client, ym)
        return {'src': src, 'kind': 'winback', 'table': 'raw_winback', 'display': '电访召回',
                'df': wdf, 'rows': len(wdf), 'is_new': True, 'months': [ym],
                'status': (f'🔄 刷新 {ym}（覆盖该月旧的电访统计）' if exists else f'🆕 新增 {ym}')}

    # 2) 客服对话（文件名带「客服对话」）
    if '客服对话' in src:
        df = _parse_cs_df(it, tmp_path, src)
        dup = _cs_basename_loaded(client, os.path.basename(src))
        return {'src': src, 'kind': 'cs', 'display': '客服对话', 'df': df,
                'rows': len(df), 'is_new': dup == 0,
                'status': ('⏭ 已上传过，会跳过' if dup else '🆕 待写入')}
    # 3) 读列判断 红利 / 客服对话（红利 zip 内层档名是纯数字，认不出，只能看列）
    ext = os.path.splitext(src)[1].lower()
    try:
        if ext in ('.xlsx', '.xls'):
            import pandas as pd
            cols = set(str(c).strip() for c in pd.read_excel(tmp_path, nrows=5).columns)
        else:
            cols = set(str(c).strip() for c in it.read_file(tmp_path).columns)
    except Exception:
        cols = set()
    ct = _commission_table(cols)
    if ct:
        df = _parse_commission(it, tmp_path)
        months = _commission_months(df)
        exists = _commission_month_exists(client, ct, months)
        disp = '代理佣金(团队版)' if ct.endswith('team') else '代理佣金(单线版)'
        mlbl = ','.join(months) if months else '?'
        return {'src': src, 'kind': 'commission', 'table': ct, 'display': disp,
                'df': df, 'rows': len(df), 'is_new': len(df) > 0, 'months': months,
                'status': (f'🔄 刷新 {mlbl}（其他月不动）' if exists > 0 else f'🆕 新增 {mlbl}')}
    if {'满意度评价', '对话内容'} & cols:
        df = _parse_cs_df(it, tmp_path, src)
        dup = _cs_basename_loaded(client, os.path.basename(src))
        return {'src': src, 'kind': 'cs', 'display': '客服对话', 'df': df,
                'rows': len(df), 'is_new': dup == 0,
                'status': ('⏭ 已上传过，会跳过' if dup else '🆕 待写入')}
    if {'订单号', '红利金额'} <= cols or {'红利标题', '申请时间'} <= cols:
        df = _parse_bonus_df(it, tmp_path)
        existing = _existing_bonus_orders(client)
        new_n = (int((~df['订单号'].astype(str).isin(existing)).sum())
                 if '订单号' in df.columns else len(df))
        return {'src': src, 'kind': 'bonus', 'display': '红利记录', 'df': df,
                'rows': len(df), 'is_new': new_n > 0, '_existing_orders': existing,
                'status': (f'🆕 待写入（新订单 {new_n} 笔）' if new_n > 0
                           else '⏭ 订单全部已存在，会跳过')}
    return {'src': src, 'kind': 'none', 'display': '⚠️ 未识别', 'df': None,
            'rows': 0, 'is_new': False, 'status': '跳过（认不出类型）'}


def _recover_zip_name(name: str, is_utf8: bool) -> str:
    """zipfile 对非 UTF-8 条目名用 cp437 解码→中文变乱码；还原回 GBK/Big5。"""
    if is_utf8:
        return name
    for enc in ('gbk', 'big5'):
        try:
            return name.encode('cp437').decode(enc)
        except Exception:
            continue
    return name


def _expand_upload_units(f, zip_password: str):
    """把一个上传档展开成若干 (source_name, tmp_path, err)。
    普通档→1 个；zip→解压里面每个 csv/xlsx（用密码 + 还原中文名）。调用方负责删 tmp_path。"""
    import os
    import tempfile
    import zipfile

    name = f.name
    ext = os.path.splitext(name)[1].lower()
    units = []
    if ext == '.zip':
        zpath = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as zt:
                zt.write(f.getbuffer())
                zpath = zt.name
            zf = zipfile.ZipFile(zpath)
            pwd = zip_password.encode() if zip_password else None
            data_members = [
                m for m in zf.infolist()
                if not m.is_dir()
                and os.path.splitext(m.filename)[1].lower() in ('.csv', '.xlsx', '.xls')
            ]
            if not data_members:
                units.append((name, None, 'zip 里没有 csv/xlsx'))
            for m in data_members:
                inner_ext = os.path.splitext(m.filename)[1].lower()
                src = os.path.basename(
                    _recover_zip_name(m.filename, bool(m.flag_bits & 0x800)))
                try:
                    data = zf.read(m, pwd=pwd)
                except RuntimeError as e:
                    units.append((src, None, f'解压失败（密码错？）：{str(e)[:40]}'))
                    continue
                with tempfile.NamedTemporaryFile(delete=False, suffix=inner_ext) as itmp:
                    itmp.write(data)
                    units.append((src, itmp.name, None))
        except Exception as e:
            units.append((name, None, f'zip 错误：{str(e)[:50]}'))
        finally:
            if zpath:
                try:
                    os.unlink(zpath)
                except Exception:
                    pass
    else:
        suffix = ext or '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(f.getbuffer())
            units.append((name, tmp.name, None))
    return units



# V6.4: render_agent_market_monthly moved to features/agent_channel.py


@st.cache_data(ttl=600)
def _data_health_rows():
    """逐表查数据最新日期 + 行数（一条 UNION 查询，省调用）。kind: auto=每日自动 / manual=手动上传。"""
    CFG = [
        ('存款订单 · 存取款分析', 'raw_finance_deposit', '完成时间', 'auto'),
        ('提款订单 · 存取款分析', 'raw_finance_withdraw', '完成时间', 'auto'),
        ('平台报表 · 经营总览', 'raw_platform_report', '日期', 'manual'),
        ('财务报表 · 经营总览', 'raw_finance_report', '时间', 'manual'),
        ('游戏报表场馆 · 游戏&场馆', 'raw_game_report_venue', '时间', 'manual'),
        ('游戏分析 · 游戏&场馆', 'raw_game_analysis', '日期', 'manual'),
        ('推广报表 · 代理&渠道', 'raw_promotion_report', '日期', 'manual'),
        ('代理报表 · 代理&渠道', 'raw_agent_report', '日期', 'manual'),
        ('即时注单 · 实时波动', 'raw_realtime_bet', '时间', 'manual'),
        ('红利记录 · 红利分析', 'raw_bonus_report', '申请时间', 'manual'),
        ('会员报表 · 会员/新注册', 'raw_member_report', '注册时间', 'manual'),
        ('TOP报表 · 会员价值', 'raw_top_report', '_snapshot_month', 'manual'),
        ('客服对话 · 客服分析', 'raw_cs_conversations', '_snapshot_month', 'manual'),
        ('电访 · 电访召回', 'raw_winback', '月份', 'manual'),
        ('代理结算汇总 · 代理佣金', 'raw_agent_settlement_summary', '月份', 'manual'),
        ('代理结算月度(市代) · 代理佣金', 'raw_agent_settlement_monthly', '月份', 'manual'),
    ]
    client = get_bq_client()
    # 先查哪些表真的存在，避免某张表缺失（沙盒过期/从未上传）让整条 UNION 查询 404 拖垮整页
    existing = None
    try:
        edf = client.query(
            f"SELECT table_name FROM `{BQ_PREFIX}`.INFORMATION_SCHEMA.TABLES"
        ).result().to_dataframe()
        existing = set(edf['table_name'].tolist())
    except Exception:
        existing = None  # 查不到就退回旧行为（全部纳入查询）
    rows = []
    parts = []
    for name, tbl, col, kind in CFG:
        if existing is not None and tbl not in existing:
            rows.append((name, kind, None, 0))  # 表不存在 → 该行单独标无数据，不拖垮其他表
            continue
        parts.append(
            f"SELECT '{name}' AS name, '{kind}' AS kind, "
            f"CAST(MAX(SUBSTR(CAST(`{col}` AS STRING),1,10)) AS STRING) AS max_d, "
            f"COUNT(*) AS n FROM `{BQ_PREFIX}.{tbl}`"
        )
    if parts:
        sql = "\nUNION ALL\n".join(parts)
        try:
            df = client.query(sql).result().to_dataframe()
            for _, r in df.iterrows():
                rows.append((r['name'], r['kind'], r['max_d'], int(r['n'])))
        except Exception as e:
            rows.append((f'查询失败：{str(e)[:80]}', 'auto', None, 0))
    # 运营日报谷歌表（自动）
    try:
        daily = load_daily_ops(_recent_month_labels(3))
        if daily is not None and not daily.empty and '日期' in daily.columns:
            md = str(daily['日期'].max())[:10]
            rows.append(('运营日报谷歌表 · 近期走势', 'auto', md, int(len(daily))))
    except Exception:
        pass
    return rows


def render_data_health():
    import datetime as _dt
    hero('数据健康',
         '一览各报表数据更新到几号、是否滞后，便于及时补数与交接。绿＝最新，黄＝稍旧，红＝需尽快更新。',
         source_badge='数据健康检查')
    today = _dt.date.today()
    cur_ym = (today.year, today.month)

    def _ym(s):
        s = str(s)
        if len(s) >= 7 and s[4] == '-':
            return (int(s[:4]), int(s[5:7]))
        if len(s) >= 6 and s[:6].isdigit():
            return (int(s[:4]), int(s[4:6]))
        return None

    def _month_gap(s):
        ym = _ym(s)
        if not ym:
            return None
        return (cur_ym[0] - ym[0]) * 12 + (cur_ym[1] - ym[1])

    out = []
    red = yellow = 0
    for name, kind, max_d, n in _data_health_rows():
        if not max_d or max_d == 'None':
            status, note = '❌ 无数据', '表为空或查询失败'
            red += 1
        elif kind == 'auto':
            try:
                d = _dt.date.fromisoformat(max_d)
                gap = (today - d).days
            except Exception:
                gap = None
            if gap is None:
                status, note = '⚠️ 待查', '日期解析失败'
                yellow += 1
            elif gap <= 1:
                status, note = '✅ 最新', '自动更新中'
            elif gap <= 3:
                status, note = f'⚠️ 滞后 {gap} 天', '检查排程是否正常'
                yellow += 1
            else:
                status, note = f'❌ 滞后 {gap} 天', '排程可能失败，需检查'
                red += 1
        else:  # manual
            mg = _month_gap(max_d)
            if mg is None:
                status, note = '⚠️ 待查', '日期解析失败'
                yellow += 1
            elif mg <= 0:
                status, note = '✅ 本月已更新', '手动上传'
            elif mg == 1:
                status, note = '⚠️ 只到上月', '本月待上传'
                yellow += 1
            else:
                status, note = f'❌ 落后 {mg} 个月', '需补上传'
                red += 1
        cat = '🤖 自动' if kind == 'auto' else '📤 手动上传'
        out.append({'报表': name, '类别': cat, '数据最新': max_d or '—',
                    '状态': status, '说明': note, '_sev': (2 if status.startswith('❌') else 1 if status.startswith('⚠️') else 0)})

    if red:
        st.error(f'有 {red} 张表需尽快更新（红）。', icon='🚨')
    elif yellow:
        st.warning(f'有 {yellow} 张表稍旧（黄），其余正常。', icon='⚠️')
    else:
        st.success('所有数据均为最新 ✅', icon='✅')

    df = pd.DataFrame(out).sort_values('_sev', ascending=False).drop(columns='_sev')
    st.dataframe(df, use_container_width=True, hide_index=True)
    st.caption('判定：自动表期望更新到昨日（滞后>3天标红）；手动表期望本月已上传（只到上月标黄、更早标红）。'
               '需要补哪些、从后台哪里导，见「数据说明」页。缓存 10 分钟，刷新页面可重查。')


def render_data_source_guide():
    hero('数据说明',
         '本页汇整面板各页的数据来源、对应后台位置与更新方式，供日常维护与交接查阅。',
         source_badge='面板数据地图')
    st.markdown(
        '#### 数据更新方式（三类）\n'
        '- **自动**：「近期走势(日报)」读取运营日报谷歌表；「存取款分析」由程序每日 11:00 抓取。两者无需人工操作。\n'
        '- **手动上传**：其余多数页面。后台导出报表后，于「数据上传」页上传；同期重传将自动覆盖，不产生重复。\n'
        '- **人工提供**：电访（电访团队提供）、客服对话（客服系统导出），取得后经上传页入库。\n'
        '  （代理结算月报、市代月度虽由客服主管/平哥提供，但现已可直接拖上传页自助入库，归手动上传。）')

    st.markdown('#### 后台报表对应导出位置')
    backend = pd.DataFrame([
        ['经营报表', '报表中心 → 经营报表', '手动上传'],
        ['财务报表', '报表中心 → 财务报表', '手动上传'],
        ['游戏报表(场馆)', '报表中心 → 游戏报表(场馆)', '手动上传'],
        ['游戏分析', '报表中心 → 游戏分析（导出选「日报」颗粒度）', '手动上传'],
        ['推广报表', '报表中心 → 推广报表', '手动上传'],
        ['代理报表', '报表中心 → 代理报表（密码 zip）', '手动上传'],
        ['会员报表', '报表中心 → 会员报表（按注册时间、完整日期、全部页数导出）', '手动上传'],
        ['TOP报表', '报表中心 → TOP报表', '手动上传'],
        ['即时注单', '报表中心 → 即时注单', '手动上传'],
        ['红利记录', '会员管理 → VIP记录管理 → 红利记录', '手动上传'],
        ['代理佣金(单线/团队)', '代理管理 → 佣金管理 → 发放佣金（设佣金月份，导出 Csv）', '手动上传'],
        ['存款/提款 历史记录', '财务管理 → 存款管理/提款管理 → 历史记录', '自动 11:00'],
        ['运营日报(平台)', '由程序每日 10:00 写入谷歌表', '自动 10:00'],
        ['代理结算月报(平哥)', '客服主管提供「X月代理帐.xlsx」→ 拖数据上传页自助入库（选月份）', '手动上传'],
        ['代理结算月度(市代)', '市代「整理资料」表导出 xlsx（每代理每月产值/挂账）', '手动上传'],
        ['客服对话', '非后台报表，客服系统导出 xlsx', '人工提供'],
        ['撥打紀錄總表(电访)', '非后台报表，电访团队提供 xlsx', '人工提供'],
    ], columns=['报表', '后台位置 / 来源', '更新方式'])
    st.dataframe(backend, use_container_width=True, hide_index=True)

    st.markdown('#### 各页面对应报表与更新方式')
    pages = pd.DataFrame([
        ['🅰️ 财务结果', '经营总览', '经营报表＋财务报表＋红利记录＋代理结算', '手动上传'],
        ['🅰️ 财务结果', '近期走势(日报)', '运营日报谷歌表「平台报表」分页', '自动（每日10:00）'],
        ['🅰️ 财务结果', '存取款分析', '存款/提款 历史记录', '自动（每日11:00）'],
        ['🅰️ 财务结果', '红利分析', '红利记录', '手动上传'],
        ['🅰️ 财务结果', '红利 ROI & 代理质量', '红利记录＋代理报表', '手动上传'],
        ['🅰️ 财务结果', '代理佣金 & 退成', '代理佣金(单线/团队)＋代理结算月报(平哥)', '手动上传'],
        ['🅱️ 会员价值', '会员结构 & ARPU', '会员报表＋TOP报表', '手动上传'],
        ['🅱️ 会员价值', '投注分析', '注单明细(按月)；来源待确认，目前仅含 4 月', '手动上传'],
        ['🅱️ 会员价值', '客服分析', '客服对话＋会员报表', '人工提供'],
        ['🅱️ 会员价值', '电访召回', '撥打紀錄總表', '人工提供'],
        ['🅱️ 会员价值', '实时波动 & DAU', '即时注单', '手动上传'],
        ['🅲 代理/渠道', '代理团队 & 渠道', '代理报表＋推广报表', '手动上传'],
        ['🅲 代理/渠道', '新注册分析', '会员报表（口径为注册数的非代理部分）', '手动上传'],
        ['🅲 代理/渠道', '代理 × 会员 明细', '会员报表＋代理报表', '手动上传'],
        ['🅲 代理/渠道', '市代月度结算', '市代「整理资料」表（每代理每月）', '手动上传'],
        ['🅲 代理/渠道', '游戏 & 场馆', '游戏报表(场馆)＋游戏分析', '手动上传'],
    ], columns=['分组', '页面', '对应报表', '更新方式'])
    st.dataframe(pages, use_container_width=True, hide_index=True)

    st.info('各页面的补数月份与口径说明，见该页顶部「数据详情」。导出注意事项：日期字段须正确、日期范围完整、导出全部页数。'
            '「投注分析」所用注单明细来源待确认（目前仅含 4 月）。')


def _validate_entry(entry, client):
    """上传完整性校验（只提示不阻断）。返回 ⚠️ 提示字符串列表。
    四项：①会员报表非代理数 vs 平台注册数 ②行数 vs 库中同期骤降 ③日期区间缺天 ④关键数值全空。"""
    warns = []
    df = entry.get('df')
    table = entry.get('table')
    if df is None or len(df) == 0:
        return warns

    # ① 会员报表完整度：最新注册月「非代理」数 vs 平台报表同月同区间「注册数」
    if table == 'raw_member_report' and '注册时间' in df.columns and '是否为代理' in df.columns:
        try:
            rt = to_datetime_safe(df['注册时间'])
            valid = rt.notna()
            if valid.any():
                latest_m = rt[valid].dt.strftime('%Y-%m').max()
                in_m = valid & (rt.dt.strftime('%Y-%m') == latest_m)
                max_day = rt[in_m].max().strftime('%Y-%m-%d')
                file_nonagent = int((in_m & (df['是否为代理'].astype(str) == '非代理')).sum())
                # 比对基准：优先用日报谷歌表（当月最新、自动），否则退 BQ 平台报表（历史月）
                plat = 0
                try:
                    daily = load_daily_ops(_recent_month_labels(3))
                    if daily is not None and not daily.empty and {'日期', '注册数'}.issubset(daily.columns):
                        dd = daily.copy()
                        ds = dd['日期'].astype(str)
                        sub = dd[(ds.str[:7] == latest_m) & (ds.str[:10] <= max_day)]
                        if len(sub):
                            plat = int(pd.to_numeric(sub['注册数'], errors='coerce').fillna(0).sum())
                except Exception:
                    plat = 0
                if plat == 0:
                    try:
                        sql = (f"SELECT SUM(SAFE_CAST(REPLACE(CAST(`注册数` AS STRING),',','') AS FLOAT64)) AS n "
                               f"FROM `{BQ_PREFIX}.raw_platform_report` "
                               f"WHERE SUBSTR(CAST(`日期` AS STRING),1,7)=@m "
                               f"AND SUBSTR(CAST(`日期` AS STRING),1,10)<=@d")
                        cfg = bigquery.QueryJobConfig(query_parameters=[
                            bigquery.ScalarQueryParameter('m', 'STRING', latest_m),
                            bigquery.ScalarQueryParameter('d', 'STRING', max_day)])
                        rows = list(client.query(sql, job_config=cfg).result())
                        plat = int(rows[0].n) if rows and rows[0].n else 0
                    except Exception:
                        plat = 0
                if plat > 0 and file_nonagent < 0.8 * plat:
                    warns.append(f'⚠️ 完整度：{latest_m} 非代理 {file_nonagent} 笔，平台同期注册 {plat} 笔'
                                 f'（仅 {file_nonagent/plat*100:.0f}%）——这份可能不完整，导出疑似带筛选或没导全部页数')
        except Exception:
            pass

    # ② 行数骤降：本次 vs 库中同期已有
    exists = entry.get('exists')
    if isinstance(exists, int) and exists >= 20 and entry['rows'] < 0.7 * exists:
        warns.append(f'⚠️ 行数偏少：本次 {entry["rows"]} 行，库中同期已有 {exists} 行'
                     f'（{entry["rows"]/exists*100:.0f}%）——是否只导了部分？')

    # ③ 日期缺口：日明细类，区间内缺天
    keys = entry.get('date_keys')
    if keys and len(keys) >= 2:
        try:
            import datetime as _dt
            ds = [_dt.date.fromisoformat(k) for k in keys if len(k) == 10 and k[4] == '-']
            if len(ds) >= 2:
                full = {ds[0] + _dt.timedelta(days=i) for i in range((max(ds) - min(ds)).days + 1)}
                miss = sorted(full - set(ds))
                if miss:
                    shown = '、'.join(d.isoformat() for d in miss[:5]) + ('…' if len(miss) > 5 else '')
                    warns.append(f'⚠️ 区间内缺 {len(miss)} 天：{shown}（导出是否跳过了某些日期？）')
        except Exception:
            pass

    # ④ 关键数值全空：常见金额列若全为 0 / 空
    for col in ('有效投注额', '公司输赢', '订单金额', '存款额', '红利金额'):
        if col in df.columns:
            num = pd.to_numeric(df[col].astype(str).str.replace(',', '', regex=False), errors='coerce')
            if num.notna().sum() == 0 or (num.fillna(0) == 0).all():
                warns.append(f'⚠️ 「{col}」整列为 0 或空——数据疑似异常，请确认导出是否正确')
            break
    return warns



def _load_dataframe_to_bq_chunked(client, df: pd.DataFrame, table_name: str, source_file: str, chunk_size: int = 50000) -> int:
    """V7 大档案上传：分批清洗、分批写入 BigQuery，降低 Streamlit RAM 峰值。"""
    if df is None or df.empty:
        return 0
    total = len(df)
    written = 0
    table_id = f"{BQ_PREFIX}.{table_name}"
    for start in range(0, total, int(chunk_size)):
        end = min(start + int(chunk_size), total)
        chunk = df.iloc[start:end].copy()
        payload = clean_upload_dataframe(chunk)
        payload['_imported_at'] = pd.Timestamp.now()
        payload['_source_file'] = source_file
        cfg = bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            autodetect=True,
            schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION],
        )
        job = client.load_table_from_dataframe(payload, table_id, job_config=cfg)
        job.result()
        written += int(job.output_rows or len(payload))
        del chunk, payload
        gc.collect()
    return written


def _write_upload_history(client, records: list[dict]) -> None:
    """写入上传历史；失败不影响主流程。"""
    if not records:
        return
    try:
        hist = pd.DataFrame(records)
        hist['_logged_at'] = pd.Timestamp.now()
        cfg = bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            autodetect=True,
            schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION],
        )
        client.load_table_from_dataframe(hist, f"{BQ_PREFIX}.upload_history", job_config=cfg).result()
    except Exception:
        pass


def _write_entry_v7(client, e: dict, src: str, chunk_size: int) -> tuple[bool, str, int]:
    """写入单一已解析 entry。返回 (ok, message, rows)。"""
    kind = e.get('kind')
    try:
        if kind == 'standard':
            # 大表优先走 append chunk；带日期替换的日报类仍保留原安全替换逻辑。
            if e.get('date_range'):
                rng, removed, written, total = _replace_by_date_range(client, e['df'], e['table'], src)
                return True, f"{e['display']}：已替换 {rng}（覆盖旧 {removed} 行、写新 {written} 行，表共 {total} 行）", int(written)
            if _bq_count_source_file(client, e['table'], src) > 0:
                return False, f"{src}：已存在，跳过", 0
            n = _load_dataframe_to_bq_chunked(client, e['df'], e['table'], src, chunk_size)
            return True, f"{e['display']}「{src}」：+{n} 行", int(n)
        if kind == 'snapshot':
            # 快照表需要同月合并，仍采用原安全逻辑。
            mth, updated, added, total = _replace_by_snapshot_month(client, e['df'], e['table'], src)
            return True, f"{e['display']}：已合并 {mth}（更新 {updated}、新增 {added}，表共 {total} 行）", int(added)
        if kind == 'bonus':
            n = _append_bonus(client, e['df'], src, e.get('_existing_orders'))
            if n == 0:
                return False, f"{src}：订单全部已存在，跳过", 0
            return True, f"红利「{src}」：+{n} 笔新订单", int(n)
        if kind == 'cs':
            if _cs_basename_loaded(client, os.path.basename(src)) > 0:
                return False, f"{src}：已存在，跳过", 0
            n = _append_cs(client, e['df'], src)
            return True, f"客服对话「{src}」：+{n} 条", int(n)
        if kind == 'winback':
            total = _write_winback(client, e['df'], e.get('months', []), src)
            return True, f"电访召回：已存 {','.join(e.get('months', []))}（表共 {total} 行）", int(total)
        if kind == 'agent_monthly':
            months = e.get('months', [])
            n = _write_agent_monthly(client, e['df'], months, src)
            rng = f"{months[0]}~{months[-1]}" if len(months) > 1 else (months[0] if months else '?')
            return True, f"代理结算月度(市代)：已按月替换 {rng}，写 {n} 行", int(n)
        if kind == 'settlement':
            return False, f"{src}：V7 串流模式暂不处理需手动选月份的代理结算月报，请切回兼容模式上传。", 0
        if kind == 'commission':
            return False, f"{src}：V7 串流模式暂不处理代理佣金整月刷新，请切回兼容模式上传。", 0
        return False, f"{src}：未识别，跳过", 0
    except Exception as ex:
        return False, f"{src}：{str(ex)[:180]}", 0


def _render_v7_stream_upload(files, zip_pw: str, chunk_size: int, auto_sync: bool) -> None:
    """V7 上传引擎：逐档处理、逐档写入、立即释放内存。"""
    it = _import_tool()
    client = get_bq_client()
    all_units = []
    for f in files:
        all_units.extend(_expand_upload_units(f, zip_pw))
    total_units = len(all_units)
    if total_units == 0:
        st.warning('没有可处理的文件。')
        return

    overall = st.progress(0.0)
    status_box = st.empty()
    results_box = st.container()
    logs: list[dict] = []
    ok_messages: list[str] = []
    fail_messages: list[str] = []
    t0 = time.time()

    for idx, (src, tmp_path, err) in enumerate(all_units, start=1):
        file_t0 = time.time()
        status_box.info(f'正在处理 {idx}/{total_units}：{src}')
        rows = 0
        ok = False
        msg = ''
        try:
            if err:
                msg = f'{src}：{err}'
                fail_messages.append(msg)
            else:
                entry = _classify_and_parse(it, client, src, tmp_path)
                rows = int(entry.get('rows') or 0)
                if not entry.get('is_new') or rows <= 0:
                    msg = f"{src}：{entry.get('status', '无需写入')}"
                    fail_messages.append(msg)
                else:
                    ok, msg, written = _write_entry_v7(client, entry, src, chunk_size)
                    rows = int(written or rows)
                    if ok:
                        ok_messages.append(msg)
                    else:
                        fail_messages.append(msg)
                # 关键：不把 df 留在 session，立即清掉。
                try:
                    if isinstance(entry, dict) and 'df' in entry:
                        del entry['df']
                    del entry
                except Exception:
                    pass
        except Exception as ex:
            msg = f'{src}：{str(ex)[:180]}'
            fail_messages.append(msg)
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
            gc.collect()

        elapsed = time.time() - file_t0
        logs.append({
            'source_file': src,
            'rows': rows,
            'seconds': round(elapsed, 2),
            'status': 'success' if ok else 'skipped_or_failed',
            'message': msg,
        })
        overall.progress(idx / total_units)
        with results_box:
            if ok:
                st.success(msg)
            else:
                st.warning(msg)

    _write_upload_history(client, logs)

    if ok_messages and auto_sync:
        with st.spinner('正在自动更新 Dashboard / Member360 / 风控中心资料表...'):
            try:
                refresh_results = refresh_core_marts(client)
                if all(r.ok for r in refresh_results):
                    st.success('核心资料表已自动更新：\n\n' + format_refresh_results(refresh_results))
                else:
                    st.warning('部分核心资料表更新失败：\n\n' + format_refresh_results(refresh_results))
            except Exception as ex:
                st.warning(f'资料已写入，但自动更新核心资料表失败：{str(ex)[:180]}。可在「同步核心资料表」手动执行。')
    try:
        st.cache_data.clear()
    except Exception:
        pass

    total_elapsed = time.time() - t0
    st.success(f'V7 上传流程完成，用时 {total_elapsed:.1f} 秒。成功 {len(ok_messages)} 个，跳过/失败 {len(fail_messages)} 个。')
    st.dataframe(pd.DataFrame(logs), use_container_width=True, hide_index=True)

def _render_data_upload_impl():
    import os

    hero('数据上传（= 把报表永久保存进数据库）',
         '这页是把数据「存起来累积」用的：把后台下载的「任何一份」月度报表直接拖进来——平台 / 财务 / 游戏 / 游戏场馆 / '
         '游戏分析 / 推广 / 代理 / 会员 / TOP / 实时注单 / 红利 / 客服对话 / 代理佣金（单线·团队）/ 电访（撥打紀錄總表）全部通吃，'
         '系统自动认出是哪张、清洗、【保存进数据库（BigQuery）永久留存】。一次拖一堆不同的、密码 zip、分好几份都行；'
         '只「新增」绝不动库里旧数据。（拖进来 = 存进数据库；只想看分析不存，去对应的分析页。）')

    it = _import_tool()
    client_for_sync = get_bq_client()
    with st.expander('🔄 同步核心资料表（raw → Dashboard / Member360 / 风控）', expanded=False):
        st.caption('上传资料后会自动同步；如果你手动修过 BigQuery，或首页最新经营日没更新，可以按这里重建核心资料表。')
        if st.button('立即同步核心资料表', key='sync_core_marts_btn'):
            with st.spinner('正在重建 fact_member_daily_v2 / mart_member_profile / risk_member_score...'):
                try:
                    sync_results = refresh_core_marts(client_for_sync)
                    if all(r.ok for r in sync_results):
                        st.success(format_refresh_results(sync_results))
                    else:
                        st.warning(format_refresh_results(sync_results))
                    try:
                        st.cache_data.clear()
                    except Exception:
                        pass
                except Exception as ex:
                    st.error(f'同步失败：{str(ex)[:220]}')

    files = st.file_uploader(
        '上传月度报表（可一次多个；支持 .xlsx / .csv / .zip）',
        type=['xlsx', 'xls', 'csv', 'zip'], accept_multiple_files=True, key='dataup_files')
    zip_pw = st.text_input(
        '压缩档解压密码（只有上传 .zip 才需要填，例如 代理 / 会员 / 推广 的密码档）',
        type='password', key='dataup_zip_pw', help='密码只在本次使用，不会保存。')

    use_v7 = st.toggle('🚀 使用 V7 大档案稳定上传引擎（建议开启）', value=True,
                       help='逐档处理、分批写入 BigQuery、自动释放内存，避免 Streamlit Cloud 上传大档时崩溃。')
    if use_v7:
        chunk_size = st.selectbox('BigQuery 分批写入大小', [10000, 20000, 50000, 100000], index=2,
                                  help='档案越大，建议选 20000 或 50000；内存较小选 10000。')
        auto_sync = st.checkbox('上传完成后自动同步核心资料表', value=True,
                                help='自动重建 fact_member_daily_v2 / mart_member_profile / risk_member_score，并清除快取。')
        if not files:
            st.info('👆 请上传 .xlsx / .csv / .zip。V7 会逐档写入，避免大档案把 Streamlit 撑爆。')
            return
        if st.button('🚀 开始 V7 稳定上传', type='primary', key='dataup_v7_start'):
            _render_v7_stream_upload(files, zip_pw, int(chunk_size), bool(auto_sync))
        else:
            st.info('已选择 V7 上传模式。确认档案无误后，按「开始 V7 稳定上传」。')
        return

    if not files:
        st.info(
            '👆 把后台下载的报表原档拖进来 = 自动保存进数据库（永久留存）。'
            '支持：平台 / 财务 / 游戏 / 游戏分析 / 推广 / 代理 / 会员 / TOP / 实时注单 / 红利 / 客服对话 / 代理佣金(单线·团队) / 电访(撥打紀錄總表)。\n\n'
            '· 文件名保持后台下载时的原样——系统靠文件名/内容认出是哪张报表，请别改名。\n\n'
            '· 密码压缩档（.zip）也能直接拖进来，在上面填一次解压密码即可，不用自己先解压。\n\n'
            '· 同一个档案重复上传会自动跳过；红利按「订单号」去重只补新订单；代理佣金按「佣金月份」刷新当月、保留其他月——都不会变成两份，也不动其他旧数据。')
        return

    client = get_bq_client()
    preview_rows = []
    parsed = []  # 每个 entry 是 _classify_and_parse 返回的 dict
    upload_warnings = []  # [(文件, [⚠️...])]
    for f in files:
        for src, tmp_path, err in _expand_upload_units(f, zip_pw):
            if err:
                preview_rows.append({'文件': src, '识别类型': '❌ 出错',
                                     '行数': 0, '状态': err, '校验': '—'})
                continue
            try:
                entry = _classify_and_parse(it, client, src, tmp_path)
                try:
                    warns = _validate_entry(entry, client)
                except Exception:
                    warns = []
                if warns:
                    upload_warnings.append((src, warns))
                preview_rows.append({'文件': src, '识别类型': entry['display'],
                                     '行数': entry['rows'], '状态': entry['status'],
                                     '校验': ('；'.join(w.replace('⚠️ ', '') for w in warns) if warns else '✅ 正常')})
                if entry['is_new'] and entry['rows'] > 0:
                    parsed.append(entry)
            except Exception as e:
                preview_rows.append({'文件': src, '识别类型': '❌ 出错',
                                     '行数': 0, '状态': str(e)[:60], '校验': '—'})
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    section_header('识别结果', '先核对识别对不对，再按下面的按钮写入。')
    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

    if upload_warnings:
        lines = []
        for src, warns in upload_warnings:
            fname = str(src).replace('\\', '/').split('/')[-1]
            for w in warns:
                lines.append(f'· **{fname}**：{w.replace("⚠️ ", "")}')
        st.warning('**完整性检查发现以下情况**（仅提示，不阻止写入；确认无误可照常写入）：\n\n' + '\n\n'.join(lines))

    if not parsed:
        st.warning('没有需要写入的新数据（都已上传过或未识别）。')
        return

    st.markdown(f'**准备写入 {len(parsed)} 个文件**（新增 / 刷新当月，不动其他旧数据）。')
    # 抓不到月份的快照报表（如 TOP 报表 top.xlsx 裸档名、无日期栏）→ 让用户选月份再写
    need_month = [e for e in parsed if e.get('need_month')]
    month_ok = True
    if need_month:
        import datetime as _dt
        _y, _mo = _dt.date.today().year, _dt.date.today().month
        _months = []
        for _ in range(13):
            _months.append(f'{_y:04d}-{_mo:02d}')
            _mo -= 1
            if _mo == 0:
                _mo = 12; _y -= 1
        opts = ['— 请选择 —'] + _months
        st.markdown('**下列报表判不出是哪个月，请选月份**（例如 TOP 报表档名只有 `top.xlsx`，没有日期）：')
        for e in need_month:
            fname = str(e['src']).replace('\\', '/').split('/')[-1]
            pick = st.selectbox(f'📄 {fname}（{e["display"]}）是哪个月？', opts, key=f'dataup_month_{fname}')
            if pick == '— 请选择 —':
                month_ok = False
            else:
                e['picked_month'] = pick   # 'YYYY-MM'
                if e.get('kind') == 'snapshot' and e.get('df') is not None:
                    ym = pick.replace('-', '')
                    e['df']['_snapshot_month'] = ym
                    snap = it.month_to_snapshot_date(ym)
                    if snap:
                        e['df']['_snapshot_date'] = snap
                    e['snapshot_month'] = ym
        if not month_ok:
            st.info('请先为上面每个报表选好月份，再写入。')
    # 守门：侦测到「整月汇总」误传(本该日明细) → 当场拦下、要二次确认，避免静默污染
    warn_entries = [e for e in parsed if e.get('warn_month_agg')]
    proceed = month_ok
    if warn_entries:
        st.warning('⚠️ 下列文件看起来是「整月汇总」而不是「日明细」（日期栏只有年月、没有具体到日）：\n\n'
                   + '\n'.join('· ' + str(e['src']).replace('\\', '/').split('/')[-1] for e in warn_entries)
                   + '\n\n这种文件导进去后，按日期范围筛选会整月对不上（跟其他月的日明细不一致）。'
                     '建议回后台改用「按日」颗粒度重新导出（档名会带完整起讫日，如 …【2026-05-01,2026-05-31】）。'
                     '真的要传汇总版才勾下面这格。')
        proceed = proceed and st.checkbox('我知道这是整月汇总、仍要写入', key='dataup_force_monthagg')
    if st.button('✅ 确认写入 BigQuery', type='primary', key='dataup_confirm', disabled=not proceed):
        import os as _os
        ok, fail = [], []
        others = [e for e in parsed if e['kind'] != 'commission']
        comm = [e for e in parsed if e['kind'] == 'commission']
        comm_tables = sorted(set(e['table'] for e in comm))
        total_steps = max(1, len(others) + len(comm_tables))
        prog = st.progress(0.0)
        done = 0
        # 非佣金：逐档 append
        for e in others:
            src, kind = e['src'], e['kind']
            try:
                if kind == 'standard':
                    if e.get('date_range'):
                        # 日报类：按日期替换（覆盖重叠日期，不重复）
                        rng, removed, written, total = _replace_by_date_range(client, e['df'], e['table'], src)
                        ok.append(f"{e['display']}：已替换 {rng}（覆盖旧 {removed} 行、写新 {written} 行，"
                                  f"表共 {total} 行，其他日期不动）")
                    elif _bq_count_source_file(client, e['table'], src) > 0:
                        fail.append(f'{src}：已存在，跳过')
                    else:
                        n = _append_standard_report(client, e['df'], e['table'], src)
                        ok.append(f"{e['display']}「{src}」：+{n} 行")
                elif kind == 'snapshot':
                    mth, updated, added, total = _replace_by_snapshot_month(client, e['df'], e['table'], src)
                    ok.append(f"{e['display']}：已合并 {mth}（按会员账号：更新 {updated} 个、新增 {added} 个，"
                              f"表共 {total} 行，其他月一行不动）")
                elif kind == 'bonus':
                    n = _append_bonus(client, e['df'], src, e.get('_existing_orders'))
                    if n == 0:
                        fail.append(f'{src}：订单全部已存在，跳过')
                    else:
                        ok.append(f"红利「{src}」：+{n} 笔新订单")
                elif kind == 'cs':
                    if _cs_basename_loaded(client, _os.path.basename(src)) > 0:
                        fail.append(f'{src}：已存在，跳过')
                    else:
                        n = _append_cs(client, e['df'], src)
                        ok.append(f"客服对话「{src}」：+{n} 条")
                elif kind == 'winback':
                    total = _write_winback(client, e['df'], e.get('months', []), src)
                    ok.append(f"电访召回：已存 {','.join(e.get('months', []))}（表共 {total} 行，其他月不动）")
                elif kind == 'agent_monthly':
                    months = e.get('months', [])
                    n = _write_agent_monthly(client, e['df'], months, src)
                    rng = f"{months[0]}~{months[-1]}" if len(months) > 1 else (months[0] if months else '?')
                    ok.append(f"代理结算月度(市代)：已按月替换 {rng}（{len(months)} 个月，写 {n} 行，其他月不动）")
                elif kind == 'settlement':
                    mth = e.get('picked_month', '?')
                    ns, nd = _write_agent_settlement(client, e.get('summary_records', []),
                                                     e.get('detail_records', []), mth, src)
                    ok.append(f"代理结算月报(平哥) {mth}：摘要 {ns} 行 + 明细 {nd} 行（按月替换，其他月不动）")
            except Exception as ex:
                fail.append(f'{src}：{str(ex)[:80]}')
            done += 1
            prog.progress(done / total_steps)
        # 佣金：按表分组，整表读改写（保留其他月份）
        for table in comm_tables:
            entries = [e for e in comm if e['table'] == table]
            disp = entries[0]['display']
            try:
                new_df = pd.concat([e['df'] for e in entries], ignore_index=True)
                srcs = ' / '.join(_os.path.basename(e['src']) for e in entries)
                months = sorted(set(m for e in entries for m in e.get('months', [])))
                total = _write_commission_safe(client, new_df, table, srcs)
                ok.append(f"{disp}：已更新 {','.join(months)}（表共 {total} 行，其他月保留）")
            except Exception as ex:
                fail.append(f'{disp}：{str(ex)[:90]}')
            done += 1
            prog.progress(done / total_steps)
        if ok:
            st.success('写入成功：\n\n' + '\n\n'.join('· ' + s for s in ok))
        if fail:
            st.error('未写入：\n\n' + '\n\n'.join('· ' + s for s in fail))

        if ok:
            with st.spinner('正在自动更新 Dashboard / Member360 / 风控中心资料表...'):
                try:
                    refresh_results = refresh_core_marts(client)
                    if all(r.ok for r in refresh_results):
                        st.success('核心资料表已自动更新：\n\n' + format_refresh_results(refresh_results))
                    else:
                        st.warning('部分核心资料表更新失败：\n\n' + format_refresh_results(refresh_results))
                except Exception as ex:
                    st.warning(f'资料已写入，但自动更新核心资料表失败：{str(ex)[:180]}。可在「同步核心资料表」手动执行。')
        try:
            st.cache_data.clear()
        except Exception:
            pass
        st.info('完成。首页、Member360、风控中心会读取最新资料；若画面未刷新，请按 Rerun 或重新整理页面。')


_DM_TABLES = {
    '会员报表': 'raw_member_report', 'TOP报表': 'raw_top_report',
    '平台报表': 'raw_platform_report', '财务报表': 'raw_finance_report',
    '游戏报表': 'raw_game_report', '游戏场馆': 'raw_game_report_venue',
    '游戏分析': 'raw_game_analysis', '推广报表': 'raw_promotion_report',
    '代理报表': 'raw_agent_report', '实时注单': 'raw_realtime_bet',
    '红利记录': 'raw_bonus_report', '客服对话': 'raw_cs_conversations',
    '代理佣金(单线)': 'raw_agent_commission_single', '代理佣金(团队)': 'raw_agent_commission_team',
    '电访召回': 'raw_winback',
    '存款订单': 'raw_finance_deposit', '提款订单': 'raw_finance_withdraw',
    '代理结算月度(市代)': 'raw_agent_settlement_monthly',
}


def render_data_manage():
    hero('删除数据（真删 BigQuery · 控制库大小）',
         '选一张表 + 要删的月份/批次 → 真的从 BigQuery 把那批删掉（表会变小、省空间），不是只是不显示。'
         '用「读出来 → 去掉 → 整张写回」的安全删法（这项目没开 billing、免费版禁直接 DELETE，但这招照样真删）。删了就没了，谨慎操作。')
    try:
        client = get_bq_client()
    except Exception as e:
        st.error(f'连不上数据库：{str(e)[:100]}')
        return
    name = st.selectbox('选一张表', list(_DM_TABLES.keys()), key='dm_table')
    table = _DM_TABLES[name]
    key, periods = _bq_periods(client, table)
    if not periods:
        st.info('这张表目前没有数据（或读取失败）。')
        return
    label = '月份' if key in ('_snapshot_month', '佣金月份') else ('上传批次（档名）' if key == '_source_file' else key)
    section_header(f'{name} 现有数据（按{label}）', f'下面是这张表里现有的{label}和各自行数，勾选要删的。')
    dfp = pd.DataFrame([{label: k, '行数': v} for k, v in periods.items()])
    st.dataframe(dfp, use_container_width=True, hide_index=True)
    sel = st.multiselect(f'选要删的{label}（可多选）', list(periods.keys()), key='dm_sel')
    if sel:
        total_del = sum(periods[s] for s in sel)
        st.warning(f'⚠️ 将从 BigQuery 真的删掉：「{name}」的 {len(sel)} 个{label}，共 {total_del} 行。'
                   f'删了无法复原（要回来只能重新上传）。')
        confirm = st.text_input('确认请输入「删除」二字', key='dm_confirm')
        if st.button('🗑 确认从 BigQuery 删除', type='primary', key='dm_btn',
                     disabled=(confirm.strip() != '删除')):
            with st.spinner('删除中（读出 → 去掉 → 写回）…'):
                try:
                    removed, remaining = _bq_delete_periods(client, table, key, sel)
                    st.success(f'✅ 已从 BigQuery 删除 {removed} 行；「{name}」现在剩 {remaining} 行（表已变小）。')
                    try:
                        st.cache_data.clear()
                    except Exception:
                        pass
                except Exception as ex:
                    st.error(f'删除失败：{str(ex)[:140]}')


